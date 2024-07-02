from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os, time

# LIST OF SCAM / FAKE TOKENS / WANT TO DELETE FOR SOME REASON
fakeTokens =['BOBA', 'SGB', 'SGB2', 'SKY2', 'VIB', 'AUDIO', 'LIT2', 'ONT', 'MDX2', 'DODO', 'SYS', 'PNT2', 'SKY', 'VIB', 'OCEAN', 'KSM', 'CGB', 'AXAXIO']

# COPY FILE TO NEW ONE
os.system('cp original.xlsx newV3.xlsx')

# FROM ORIGINAL
wb = load_workbook('newV3.xlsx')
ws = wb.active

# NUM TX
print(len(ws['A'])-2, ' transactions loaded')


# FILL ROWS COLORS
def printRow(row, color):

    if color == 'red':
        for i in range(65,76):
            ws[chr(i)+str(row)].fill = PatternFill("solid", start_color="FF0000")
    if color == 'green':
        for i in range(65,76):
            ws[chr(i)+str(row)].fill = PatternFill("solid", start_color="00FF00")


# CHECK IF ROW IS COLORED
def check_color(row):
    if (ws['A'+str(row)].fill.start_color.index == "00000000"):
        return True
    else:
        return False

def check_red_color(row):
    if (ws['A'+str(row)].fill.start_color.index == "00FF0000"):
        return True

B = 0
D = 0
F = 0
group = []
flag_first_in_group = True

# DETECT SAME TIME OPERATIONS
row = 3
while True:
    if (ws['A'+str(row)].value == None):
        break
    if check_color(row):
        c = 0
        mismaFecha = True
        while mismaFecha:
            c += 1 
            #print()
            #print("********")
            #print("Comparando " + str(row) + " con " + str(row + c))
            try:
                if (ws['K'+str(row)].value[0:10] != ws['K'+str(row+c)].value[0:10]):    ## MODO AGRESIVO [0:10], MODO NORMAL [0:16] 
                    #print(ws['K'+str(row)].value[0:10])
                    mismaFecha = False
                    break
            except:
                break
            #time.sleep(2)   
            if (ws['C'+str(row)].value == ws['C'+str(row+c)].value) \
                and (ws['E'+str(row)].value == ws['E'+str(row+c)].value) \
                and (ws['H'+str(row)].value == ws['H'+str(row+c)].value) \
                and (ws['A'+str(row)].value == 'Operación' or ws['A'+str(row)].value == 'Otras comisiones' or ws['A'+str(row)].value == 'Recompensa / Bonificación'):
                #print("IGUALES")
                if flag_first_in_group:
                    group.append(row)
                    printRow(row, 'red')
                    flag_first_in_group = False
                    try:
                        B += ws['B'+str(row)].value
                    except:
                        B = None
                    try:
                        D += ws['D'+str(row)].value
                        #print(D)
                    except:
                        D = None
                    try:
                        F += ws['F'+str(row)].value
                    except:
                        F = None
                    
                group.append(row+c)
                printRow(row+c, 'red')
                # SUM
                try:
                    B += ws['B'+str(row+c)].value
                except:
                    B = None
                try:
                    D += ws['D'+str(row+c)].value
                    #print("sumamos", ws['D'+str(row+c)].value)
                    #print(D)
                except:
                    D = None
                try:
                    F += ws['F'+str(row+c)].value
                except:
                    F = None 
            
        # DISTINTA FECHA
        if len(group)>0:
            #print("--> agrupacion")
            #print(group)
            flag_first_in_group = True
            # INSERT NEW ROW
            ws.insert_rows(group[-1]+1)
            ws['A'+str(group[-1]+1)] = ws['A'+str(group[-1])].value
            ws['B'+str(group[-1]+1)] = B
            ws['C'+str(group[-1]+1)] = ws['C'+str(group[-1])].value
            ws['D'+str(group[-1]+1)] = D
            ws['E'+str(group[-1]+1)] = ws['E'+str(group[-1])].value
            ws['F'+str(group[-1]+1)] = F
            ws['G'+str(group[-1]+1)] = ws['G'+str(group[-1])].value
            ws['H'+str(group[-1]+1)] = ws['H'+str(group[-1])].value
            ws['I'+str(group[-1]+1)] = ws['I'+str(group[-1])].value
            ws['J'+str(group[-1]+1)] = ws['J'+str(group[-1])].value
            ws['K'+str(group[-1]+1)] = ws['K'+str(group[-1])].value
            printRow(group[-1]+1, 'green')
            B = 0
            D = 0
            F = 0
            group = []
    
    row += 1

wb.save("changesV3.xlsx")

print("Hasta borrar", len(ws['A']) + 1)
for row in reversed(range(3, len(ws['A']) + 1)):
    if check_red_color(row):
        ws.delete_rows(row)



print(len(ws['A'])-2, ' transactions now')
wb.save("newV3.xlsx")
wb.close()
